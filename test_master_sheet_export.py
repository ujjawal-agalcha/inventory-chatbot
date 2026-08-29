import io
import sys
import unittest
import openpyxl
from fastapi.testclient import TestClient

from app import app
from auth import create_access_token
from models import SessionLocal, User
from mongo_db import get_products_collection
from services.mongo_inventory_service import (
    ensure_permanent_electronic_inventory,
    update_product_stock,
)
from services.excel_export_service import (
    generate_master_sheet_bytes,
    STOCK_RED_FILL,
    STOCK_YELLOW_FILL,
    STOCK_GREEN_FILL,
    get_stock_style,
)


class TestMasterSheetExport(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        # Ensure permanent inventory exists
        ensure_permanent_electronic_inventory()
        cls.client = TestClient(app)
        
        # Ensure a test user exists
        db = SessionLocal()
        user = db.query(User).filter(User.username == "admin").first()
        if not user:
            user = User(
                username="admin",
                email="admin@test.local",
                name="Admin User",
                hashed_password="dummy_password_hash",
            )
            db.add(user)
            db.commit()
            db.refresh(user)
        
        cls.token = create_access_token({
            "sub": user.id,
            "username": user.username,
            "name": user.name,
        })
        db.close()

    def test_01_stock_threshold_styles(self):
        """Test get_stock_style logic for all required thresholds."""
        # 1. Stock = 5  (< 15) -> RED
        fill_5, font_5 = get_stock_style(5)
        self.assertEqual(fill_5.start_color.rgb, STOCK_RED_FILL.start_color.rgb)
        self.assertEqual(fill_5.start_color.rgb, "FFFFC7CE")
        print("[PASS] Stock = 5  -> RED (Color: " + str(fill_5.start_color.rgb) + ")")

        # 2. Stock = 14 (< 15) -> RED
        fill_14, font_14 = get_stock_style(14)
        self.assertEqual(fill_14.start_color.rgb, STOCK_RED_FILL.start_color.rgb)
        self.assertEqual(fill_14.start_color.rgb, "FFFFC7CE")
        print("[PASS] Stock = 14 -> RED (Color: " + str(fill_14.start_color.rgb) + ")")

        # 3. Stock = 15 (= 15) -> YELLOW
        fill_15, font_15 = get_stock_style(15)
        self.assertEqual(fill_15.start_color.rgb, STOCK_YELLOW_FILL.start_color.rgb)
        self.assertEqual(fill_15.start_color.rgb, "FFFFEB9C")
        print("[PASS] Stock = 15 -> YELLOW (Color: " + str(fill_15.start_color.rgb) + ")")

        # 4. Stock = 16 (> 15) -> GREEN
        fill_16, font_16 = get_stock_style(16)
        self.assertEqual(fill_16.start_color.rgb, STOCK_GREEN_FILL.start_color.rgb)
        self.assertEqual(fill_16.start_color.rgb, "FFC6EFCE")
        print("[PASS] Stock = 16 -> GREEN (Color: " + str(fill_16.start_color.rgb) + ")")

        # 5. Stock = 50 (> 15) -> GREEN
        fill_50, font_50 = get_stock_style(50)
        self.assertEqual(fill_50.start_color.rgb, STOCK_GREEN_FILL.start_color.rgb)
        self.assertEqual(fill_50.start_color.rgb, "FFC6EFCE")
        print("[PASS] Stock = 50 -> GREEN (Color: " + str(fill_50.start_color.rgb) + ")")

    def test_02_master_sheet_generation_and_formatting(self):
        """Test generating Excel workbook from live MongoDB data and inspect cells."""
        excel_bytes = generate_master_sheet_bytes()
        self.assertGreater(len(excel_bytes), 1000)

        # Load workbook with openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        self.assertIn("Master Inventory", wb.sheetnames)
        ws = wb["Master Inventory"]

        # Verify Header Row
        headers = [str(cell.value) for cell in ws[1]]
        expected_headers = [
            "Component / Product Name",
            "Category",
            "Price (\u20b9)",
            "Stock Left",
            "Minimum Stock Level",
            "Supplier",
            "Status",
            "Last Updated",
        ]
        self.assertEqual(headers, expected_headers)
        print("[PASS] Headers verified in generated Excel: " + str(len(headers)) + " columns.")

        # Verify freeze panes and auto-filter
        self.assertEqual(ws.freeze_panes, "A2")
        self.assertIsNotNone(ws.auto_filter.ref)

        # Verify row contents and stock cell color
        col = get_products_collection()
        mongo_products_count = col.count_documents({})
        self.assertGreater(mongo_products_count, 0)
        self.assertEqual(ws.max_row, mongo_products_count + 1)

        print("[PASS] Verified " + str(ws.max_row - 1) + " products in Excel matching MongoDB count " + str(mongo_products_count) + ".")

        # Check each row's stock value and its formatting
        for row in range(2, ws.max_row + 1):
            prod_name = ws.cell(row=row, column=1).value
            price = ws.cell(row=row, column=3).value
            stock_left = ws.cell(row=row, column=4).value
            stock_cell = ws.cell(row=row, column=4)

            self.assertIsNotNone(prod_name)
            self.assertIsInstance(price, (int, float))
            self.assertIsInstance(stock_left, int)

            cell_fill_rgb = str(stock_cell.fill.start_color.rgb)
            if stock_left < 15:
                self.assertEqual(cell_fill_rgb, "FFFFC7CE", f"Expected RED for stock {stock_left} in row {row}")
            elif stock_left == 15:
                self.assertEqual(cell_fill_rgb, "FFFFEB9C", f"Expected YELLOW for stock {stock_left} in row {row}")
            else:
                self.assertEqual(cell_fill_rgb, "FFC6EFCE", f"Expected GREEN for stock {stock_left} in row {row}")

        print("[PASS] Verified all row values and cell fill colors match the <15 / =15 / >15 rules.")

    def test_03_no_stale_data_live_update(self):
        """Test that changing stock in MongoDB is immediately reflected with updated color."""
        # Find an item with stock 20 (e.g. Arduino Uno R3)
        prod = get_products_collection().find_one({"name": "Arduino Uno R3"})
        self.assertIsNotNone(prod, "Arduino Uno R3 not found")

        # Step 1: Set stock to 20 (> 15 -> GREEN)
        update_product_stock("Arduino Uno R3", 20)
        excel_bytes_1 = generate_master_sheet_bytes()
        wb_1 = openpyxl.load_workbook(io.BytesIO(excel_bytes_1))
        ws_1 = wb_1["Master Inventory"]

        row_num = None
        for r in range(2, ws_1.max_row + 1):
            if ws_1.cell(row=r, column=1).value == "Arduino Uno R3":
                row_num = r
                break
        
        self.assertIsNotNone(row_num)
        self.assertEqual(ws_1.cell(row=row_num, column=4).value, 20)
        self.assertEqual(ws_1.cell(row=row_num, column=4).fill.start_color.rgb, "FFC6EFCE") # GREEN
        print("[PASS] Initial state: Stock = 20 -> GREEN (FFC6EFCE)")

        # Step 2: Change stock to 12 (< 15 -> RED)
        update_product_stock("Arduino Uno R3", 12)
        excel_bytes_2 = generate_master_sheet_bytes()
        wb_2 = openpyxl.load_workbook(io.BytesIO(excel_bytes_2))
        ws_2 = wb_2["Master Inventory"]

        self.assertEqual(ws_2.cell(row=row_num, column=4).value, 12)
        self.assertEqual(ws_2.cell(row=row_num, column=4).fill.start_color.rgb, "FFFFC7CE") # RED
        print("[PASS] Dynamic update: Stock changed 20 -> 12 -> RED (FFFFC7CE). NO STALE DATA confirmed!")

        # Step 3: Change stock to 15 (= 15 -> YELLOW)
        update_product_stock("Arduino Uno R3", 15)
        excel_bytes_3 = generate_master_sheet_bytes()
        wb_3 = openpyxl.load_workbook(io.BytesIO(excel_bytes_3))
        ws_3 = wb_3["Master Inventory"]

        self.assertEqual(ws_3.cell(row=row_num, column=4).value, 15)
        self.assertEqual(ws_3.cell(row=row_num, column=4).fill.start_color.rgb, "FFFFEB9C") # YELLOW
        print("[PASS] Dynamic update: Stock changed 12 -> 15 -> YELLOW (FFFFEB9C).")

    def test_04_api_endpoint_download(self):
        """Test GET /api/inventory/master-sheet/download authentication and response."""
        # 1. Unauthenticated request without token or cookie -> 401
        res_unauth = self.client.get("/api/inventory/master-sheet/download")
        self.assertEqual(res_unauth.status_code, 401)
        print("[PASS] Unauthenticated request properly rejected with 401 Unauthorized.")

        # 2. Authenticated request with Bearer token
        res_auth = self.client.get(
            "/api/inventory/master-sheet/download",
            headers={"Authorization": f"Bearer {self.token}"}
        )
        self.assertEqual(res_auth.status_code, 200)
        self.assertEqual(
            res_auth.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        self.assertIn("attachment; filename=", res_auth.headers["content-disposition"])
        self.assertIn("Master_Inventory_Sheet_", res_auth.headers["content-disposition"])

        # 3. Verify downloaded bytes form a valid openpyxl workbook
        wb = openpyxl.load_workbook(io.BytesIO(res_auth.content))
        self.assertIn("Master Inventory", wb.sheetnames)
        print("[PASS] Authenticated endpoint returned valid .xlsx file with Content-Disposition headers.")


if __name__ == "__main__":
    suite = unittest.TestLoader().loadTestsFromTestCase(TestMasterSheetExport)
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)
    sys.exit(0 if result.wasSuccessful() else 1)
