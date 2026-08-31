import os
import re
import io
import uuid
import hashlib
import logging
from datetime import datetime
from typing import Dict, Any, List, Tuple, Optional
import openpyxl
# pyrefly: ignore [missing-import]
from bson import ObjectId

from mongo_db import (
    get_products_collection,
    get_procurement_collection,
    get_expenses_collection,
    get_imports_collection,
    init_mongo_indexes,
)

logger = logging.getLogger("excel_import")

# ============================================================
# STOP WORDS FOR KEYWORD EXTRACTION
# ============================================================
STOP_WORDS = {
    "a", "an", "and", "are", "as", "at", "be", "by", "for", "from",
    "has", "he", "in", "is", "it", "its", "of", "on", "that", "the",
    "to", "was", "were", "will", "with", "needed", "urgent", "office",
    "team", "all", "monthly", "room", "desk", "new", "old", "set", "pack",
    "s", "no", "sno", "item", "items", "details", "remarks", "price", "amount"
}

# ============================================================
# CATEGORY INFERENCE RULES
# ============================================================
CATEGORY_KEYWORDS = {
    "ESP Modules": ["esp32", "esp8266", "nodemcu", "devkit", "wroom", "esp-32"],
    "Arduino Boards": ["arduino", "uno", "nano", "mega", "leonardo"],
    "Motor Drivers": ["driver", "l298n", "bts7960", "tb6612"],
    "Motors": ["motor", "servo", "sg90", "mg996r", "stepper", "nema", "gear motor"],
    "Sensors": ["sensor", "ultrasonic", "hc-sr04", "pir", "dht11", "dht22", "mpu6050", "gyroscope", "ir obstacle"],
    "Batteries": ["battery", "18650", "lipo", "rechargeable", "cell"],
    "Displays": ["display", "lcd", "oled", "tft", "screen"],
    "Relays": ["relay", "channel relay"],
    "Communication": ["bluetooth", "hc-05", "gsm", "sim800l", "gps", "neo-6m", "lora", "zigbee", "rfid"],
    "Components": ["screw", "nut", "bolt", "resistor", "capacitor", "led", "breadboard", "jumper wire"],
    "Paper & Stationery": ["paper", "ream", "notebook", "notes", "sticky", "pen", "pencil", "marker", "whiteboard", "copier", "stationery"],
    "Office Equipment": ["stapler", "punch", "calculator", "cutter", "shredder", "laminator", "board", "extension", "socket"],
    "IT & Electronics": ["cartridge", "toner", "ink", "printer", "hdmi", "cable", "usb", "adapter", "mouse", "keyboard", "projector", "laptop", "monitor"],
    "Health & Hygiene": ["sanitiser", "sanitizer", "dettol", "soap", "mask", "disinfectant", "tissue", "cleaning"],
}


def normalize_text(text: Any) -> str:
    """Normalize string: lowercase, strip punctuation, collapse whitespace."""
    if text is None:
        return ""
    s = str(text).strip().lower()
    s = re.sub(r"[^\w\s-]", " ", s)
    return " ".join(s.split())


def clean_display_name(raw_name: Any) -> str:
    """Clean raw product name for presentation."""
    if not raw_name:
        return "Unknown Item"
    s = str(raw_name).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def parse_numeric(val: Any, default: float = 0.0) -> float:
    """Safely convert strings, currency symbols, floats to float."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    s = re.sub(r"[₹$,\s]", "", s)
    try:
        return float(s)
    except (ValueError, TypeError):
        return default


def parse_int_qty(val: Any, default: int = 0) -> int:
    """Safely convert quantity to int."""
    num = parse_numeric(val, float(default))
    return int(round(num))


def parse_date_value(val: Any) -> Tuple[datetime, str]:
    """Parse date cell into datetime object and ISO string."""
    if val is None or val == "":
        now = datetime.utcnow()
        return now, now.strftime("%Y-%m-%d")
    if isinstance(val, datetime):
        return val, val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt, dt.strftime("%Y-%m-%d")
        except ValueError:
            pass
    now = datetime.utcnow()
    return now, now.strftime("%Y-%m-%d")


def extract_keywords_and_aliases(name: str, details: str = "", remarks: str = "") -> Tuple[List[str], List[str]]:
    """
    Extract meaningful search keywords and query aliases from product attributes.
    """
    combined = f"{name} {details} {remarks}".lower()
    clean_str = re.sub(r"[^\w\s-]", " ", combined)
    raw_tokens = clean_str.split()
    
    keywords = set()
    for token in raw_tokens:
        token = token.strip("-").strip()
        if len(token) >= 2 and token not in STOP_WORDS and not token.isdigit():
            keywords.add(token)
        elif token.isdigit() and len(token) <= 5:
            keywords.add(token)

    # Aliases generation
    aliases = set()
    norm_name = normalize_text(name)
    if norm_name:
        aliases.add(norm_name)
    
    # Remove parenthetical variations
    name_clean = re.sub(r"\(.*?\)", "", name).strip()
    if name_clean and normalize_text(name_clean) != norm_name:
        aliases.add(normalize_text(name_clean))

    if details:
        norm_details = normalize_text(details)
        if norm_details:
            aliases.add(norm_details)
            aliases.add(f"{norm_name} {norm_details}")

    return sorted(list(keywords)), sorted(list(aliases))


def infer_category(name: str, details: str = "") -> str:
    """Infer category from product name and details."""
    text = f"{name} {details}".lower()
    for cat, kws in CATEGORY_KEYWORDS.items():
        for kw in kws:
            if kw in text:
                return cat
    return "General Supplies"


def extract_category_and_subcategory(
    filename: str = "",
    sheet_name: str = "",
    item_name: str = "",
    details: str = "",
    row_category: Optional[str] = None,
    row_subcategory: Optional[str] = None,
) -> Tuple[str, str]:
    """
    Extract normalized Category and Sub-category from filename, sheet name, row data, or inferred keywords.
    E.g.
    'Procurement - Nuts & Bolts.xlsx' -> Category: 'Procurement', Sub-category: 'Nuts & Bolts'
    'Procurement - Electronic Equipment.xlsx' -> Category: 'Procurement', Sub-category: 'Electronic Equipment'
    'Expenses.xlsx' -> Category: 'Expenses', Sub-category: inferred
    """
    cat = str(row_category).strip() if row_category and str(row_category).strip() else ""
    subcat = str(row_subcategory).strip() if row_subcategory and str(row_subcategory).strip() else ""

    # Clean filename without extension
    clean_fn = re.sub(r"\.xlsx?$", "", filename, flags=re.I).strip()

    # Check for "Category - Subcategory" or "Category – Subcategory" format in filename
    for sep in [" - ", " – ", " — "]:
        if sep in clean_fn:
            parts = clean_fn.split(sep, 1)
            if not cat:
                cat = parts[0].strip()
            if not subcat:
                subcat = parts[1].strip()
            break

    # Check sheet name
    if sheet_name and not is_dummy_sheet(sheet_name, []):
        clean_sn = sheet_name.strip()
        for sep in [" - ", " – ", " — "]:
            if sep in clean_sn:
                sp = clean_sn.split(sep, 1)
                if not cat:
                    cat = sp[0].strip()
                if not subcat:
                    subcat = sp[1].strip()
                break
        if not subcat and clean_sn.lower() not in ("sheet1", "sheet2", "sheet", "data", "master"):
            subcat = clean_sn

    # If still missing category, infer from product name & details
    if not cat:
        inferred = infer_category(item_name, details)
        cat = inferred if inferred != "General Supplies" else "General"

    if not subcat:
        inferred = infer_category(item_name, details)
        subcat = inferred if inferred != "General Supplies" else "General"

    return cat, subcat


def compute_row_hash(*fields) -> str:
    """Generate deterministic SHA256 hash for row deduplication."""
    content = "|".join(normalize_text(f) for f in fields)
    return hashlib.sha256(content.encode("utf-8")).hexdigest()


# ============================================================
# EXCEL SHEET INSPECTION & EXTRACTION
# ============================================================

def is_dummy_sheet(sheet_name: str, sheet_values: List[List[Any]]) -> bool:
    """Detect if sheet is a dummy/sample/overview tab to ignore."""
    name_lower = sheet_name.lower()
    dummy_names = {"overview", "dashboard", "vendor list", "vendor contacts", "annual budget", "instructions", "sheet3"}
    if name_lower in dummy_names:
        return True
    
    # Check if first few rows have dummy warnings
    for row in sheet_values[:4]:
        for cell in row:
            if cell and ("not connected to the tracker" in str(cell).lower() or "dummy tab" in str(cell).lower()):
                return True
    return False


def find_header_row(rows: List[List[Any]]) -> Tuple[int, Dict[str, int]]:
    """
    Find the header row index and map standard column names to indices.
    """
    for idx, row in enumerate(rows[:6]):
        header_map = {}
        for col_idx, cell in enumerate(row):
            if not cell:
                continue
            c_str = normalize_text(cell)
            if any(k in c_str for k in ["s no", "sno", "name", "component", "components", "product", "item", "qty", "quantity", "amount", "price", "unit price", "date", "status", "category", "supplier", "vendor"]):
                header_map[c_str] = col_idx

        if len(header_map) >= 2:
            return idx, header_map

    return 0, {}


def resolve_column_indices(col_map: Dict[str, int]) -> Dict[str, Optional[int]]:
    """Resolve column indices for any flexible variations in column headers."""
    resolved = {
        "id": None,
        "name": None,
        "category": None,
        "sub_category": None,
        "details": None,
        "qty": None,
        "amount": None,
        "price": None,
        "min_stock": None,
        "market": None,
        "date": None,
        "status": None,
        "vendor": None,
        "approved_by": None,
        "remarks": None,
        "issued_by": None,
        "url": None,
    }

    for col_name, idx in col_map.items():
        cn = col_name.lower().strip()
        
        # Product ID
        if any(k in cn for k in ["product id", "product_id", "prod id", "item id", "sku", "part number", "part no", "code"]) and not any(k in cn for k in ["name", "vendor", "supplier"]):
            if resolved["id"] is None: resolved["id"] = idx
        # Name
        elif any(k in cn for k in ["product name", "component name", "item name", "product", "component", "components", "item", "items", "part name", "description"]) and not any(k in cn for k in ["vendor", "supplier", "approv"]):
            if resolved["name"] is None: resolved["name"] = idx
        # Sub-category
        elif any(k in cn for k in ["sub category", "subcategory", "sub-category"]):
            if resolved["sub_category"] is None: resolved["sub_category"] = idx
        # Category
        elif any(k in cn for k in ["category", "division", "section", "type"]):
            if resolved["category"] is None: resolved["category"] = idx
        # Min Stock
        elif any(k in cn for k in ["minimum stock", "min stock", "min level", "threshold", "reorder level"]):
            if resolved["min_stock"] is None: resolved["min_stock"] = idx
        # Quantity
        elif any(k in cn for k in ["qty", "quantity", "units", "count", "pieces", "pcs", "stock", "current stock", "balance"]):
            if resolved["qty"] is None: resolved["qty"] = idx
        # Amount
        elif any(k in cn for k in ["amount", "total amount", "total cost", "total price", "expense", "total"]):
            if resolved["amount"] is None: resolved["amount"] = idx
        # Unit Price
        elif any(k in cn for k in ["unit price", "price", "cost", "rate", "price per unit"]):
            if resolved["price"] is None: resolved["price"] = idx
        # Market
        elif any(k in cn for k in ["market", "store"]):
            if resolved["market"] is None: resolved["market"] = idx
        # Date
        elif any(k in cn for k in ["date", "order date", "purchase date", "invoice date"]):
            if resolved["date"] is None: resolved["date"] = idx
        # Status
        elif any(k in cn for k in ["status", "order status", "state"]):
            if resolved["status"] is None: resolved["status"] = idx
        # Vendor / Supplier
        elif any(k in cn for k in ["vendor", "supplier", "dealer", "distributor"]):
            if resolved["vendor"] is None: resolved["vendor"] = idx
        # Approved by
        elif any(k in cn for k in ["approved by", "approver", "approved"]):
            if resolved["approved_by"] is None: resolved["approved_by"] = idx
        # Remarks
        elif any(k in cn for k in ["remark", "remarks", "notes", "comment"]):
            if resolved["remarks"] is None: resolved["remarks"] = idx
        # Details
        elif any(k in cn for k in ["detail", "details", "spec", "specs", "specification"]):
            if resolved["details"] is None: resolved["details"] = idx
        # Issued by
        elif any(k in cn for k in ["issued by", "requested by", "requester"]):
            if resolved["issued_by"] is None: resolved["issued_by"] = idx
        # URL
        elif "url" in cn or "link" in cn:
            if resolved["url"] is None: resolved["url"] = idx

    return resolved


def read_excel_sheets(file_bytes: bytes) -> Dict[str, List[List[Any]]]:
    """Read all sheets from Excel file bytes using openpyxl."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheets_data = {}
    for name in wb.sheetnames:
        ws = wb[name]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        sheets_data[name] = data
    return sheets_data


# ============================================================
# IMPORT PIPELINE IMPLEMENTATION (IDEMPOTENT + BATCH METADATA)
# ============================================================

def process_procurement_data(
    file_bytes: bytes,
    filename: str,
    uploaded_by: str = "admin"
) -> Dict[str, Any]:
    """
    Process Procurement / Requirements Workbook.
    Assigns unique import_batch_id, retains metadata, ensures idempotency.
    """
    init_mongo_indexes()
    products_col = get_products_collection()
    procurement_col = get_procurement_collection()
    imports_col = get_imports_collection()

    sheets = read_excel_sheets(file_bytes)
    
    import_batch_id = f"batch_proc_{uuid.uuid4().hex[:12]}"
    
    import_record = {
        "import_batch_id": import_batch_id,
        "filename": filename,
        "file_type": "procurement",
        "source_type": "excel_import",
        "upload_timestamp": datetime.utcnow(),
        "uploaded_by": uploaded_by,
        "total_rows": 0,
        "valid_records": 0,
        "new_records": 0,
        "updated_records": 0,
        "duplicate_records": 0,
        "rejected_records": 0,
        "errors": [],
        "status": "in_progress",
    }
    import_id = imports_col.insert_one(import_record).inserted_id

    total_rows = 0
    new_products_cnt = 0
    updated_products_cnt = 0
    duplicate_rows_cnt = 0
    valid_records_cnt = 0
    errors = []

    try:
        for sheet_name, rows in sheets.items():
            if is_dummy_sheet(sheet_name, rows):
                logger.info("Skipping dummy tab in procurement workbook: %s", sheet_name)
                continue

            if not rows or len(rows) < 2:
                continue

            header_idx, col_map = find_header_row(rows)
            if not col_map:
                continue

            # Resolve all column positions flexibly
            cols = resolve_column_indices(col_map)
            name_idx = cols["name"]
            id_idx = cols["id"]
            cat_idx = cols["category"]
            subcat_idx = cols["sub_category"]
            details_idx = cols["details"]
            qty_idx = cols["qty"]
            amount_idx = cols["amount"]
            price_idx = cols["price"]
            min_stock_idx = cols["min_stock"]
            market_idx = cols["market"]
            order_date_idx = cols["date"]
            order_status_idx = cols["status"]
            vendor_idx = cols["vendor"]
            approved_by_idx = cols["approved_by"]
            remarks_idx = cols["remarks"]
            issued_by_idx = cols["issued_by"]
            url_idx = cols["url"]

            # Process data rows
            for r_idx in range(header_idx + 1, len(rows)):
                row = rows[r_idx]
                if not row or not any(row):
                    continue

                total_rows += 1
                raw_name = row[name_idx] if name_idx is not None and name_idx < len(row) else None
                if not raw_name or str(raw_name).strip() == "":
                    continue

                clean_name = clean_display_name(raw_name)
                norm_name = normalize_text(clean_name)
                if not norm_name:
                    continue

                raw_id = str(row[id_idx]).strip() if id_idx is not None and id_idx < len(row) and row[id_idx] is not None else ""
                row_cat = str(row[cat_idx]).strip() if cat_idx is not None and cat_idx < len(row) and row[cat_idx] is not None else None
                row_subcat = str(row[subcat_idx]).strip() if subcat_idx is not None and subcat_idx < len(row) and row[subcat_idx] is not None else None

                details = str(row[details_idx]).strip() if details_idx is not None and details_idx < len(row) and row[details_idx] is not None else ""
                qty = parse_int_qty(row[qty_idx] if qty_idx is not None and qty_idx < len(row) else 1, default=1)
                unit_price = parse_numeric(row[price_idx] if price_idx is not None and price_idx < len(row) else 0.0)
                amount = parse_numeric(row[amount_idx] if amount_idx is not None and amount_idx < len(row) else (qty * unit_price))
                min_stock_val = parse_int_qty(row[min_stock_idx] if min_stock_idx is not None and min_stock_idx < len(row) else max(3, int(qty * 0.5)), default=max(3, int(qty * 0.5)))
                market = str(row[market_idx]).strip() if market_idx is not None and market_idx < len(row) and row[market_idx] is not None else "General"
                
                raw_date = row[order_date_idx] if order_date_idx is not None and order_date_idx < len(row) else None
                dt_obj, date_str = parse_date_value(raw_date)

                order_status = str(row[order_status_idx]).strip() if order_status_idx is not None and order_status_idx < len(row) and row[order_status_idx] is not None else "Pending"
                vendor = str(row[vendor_idx]).strip() if vendor_idx is not None and vendor_idx < len(row) and row[vendor_idx] is not None else "Standard Vendor"
                approved_by = str(row[approved_by_idx]).strip() if approved_by_idx is not None and approved_by_idx < len(row) and row[approved_by_idx] is not None else ""
                remarks = str(row[remarks_idx]).strip() if remarks_idx is not None and remarks_idx < len(row) and row[remarks_idx] is not None else ""
                issued_by = str(row[issued_by_idx]).strip() if issued_by_idx is not None and issued_by_idx < len(row) and row[issued_by_idx] is not None else ""
                url = str(row[url_idx]).strip() if url_idx is not None and url_idx < len(row) and row[url_idx] is not None else ""

                category, sub_category = extract_category_and_subcategory(
                    filename=filename,
                    sheet_name=sheet_name,
                    item_name=clean_name,
                    details=details or remarks,
                    row_category=row_cat,
                    row_subcategory=row_subcat,
                )

                # Business key deduplication hash (Idempotent across re-uploads)
                row_hash = compute_row_hash(norm_name, qty, amount, date_str, vendor, order_status, remarks, sheet_name)

                # Check if this exact procurement row was already imported
                existing_proc = procurement_col.find_one({"row_hash": row_hash})
                if existing_proc:
                    duplicate_rows_cnt += 1
                    logger.debug("Duplicate procurement row skipped: %s (%s)", clean_name, row_hash)
                    continue

                valid_records_cnt += 1
                keywords, aliases = extract_keywords_and_aliases(clean_name, details, remarks)

                # Upsert into Master Inventory (products collection)
                existing_prod = products_col.find_one({"normalized_name": norm_name})
                
                if existing_prod:
                    updated_products_cnt += 1
                    prod_id = existing_prod["_id"]
                    
                    merged_keywords = sorted(list(set(existing_prod.get("keywords", []) + keywords)))
                    merged_aliases = sorted(list(set(existing_prod.get("aliases", []) + aliases)))
                    merged_sources = list(set(existing_prod.get("source_files", []) + [filename]))
                    merged_batches = list(set(existing_prod.get("import_batch_ids", []) + [import_batch_id]))

                    stock_delta = qty if order_status.lower() == "fulfilled" else 0
                    pending_delta = qty if order_status.lower() in ("pending", "approved") else 0

                    update_fields = {
                        "updated_at": datetime.utcnow(),
                        "keywords": merged_keywords,
                        "aliases": merged_aliases,
                        "source_files": merged_sources,
                        "import_batch_ids": merged_batches,
                    }
                    if category and (not existing_prod.get("category") or existing_prod.get("category") in ("General", "General Supplies")):
                        update_fields["category"] = category
                    if sub_category and not existing_prod.get("sub_category"):
                        update_fields["sub_category"] = sub_category
                    if details and not existing_prod.get("details"):
                        update_fields["details"] = details
                    if vendor and (not existing_prod.get("supplier") or existing_prod.get("supplier") in ("Standard Vendor", "Corporate Vendor")):
                        update_fields["supplier"] = vendor
                    if market:
                        update_fields["market"] = market
                    if unit_price > 0 and (not existing_prod.get("unit_price") or existing_prod.get("unit_price") == 0.0):
                        update_fields["unit_price"] = unit_price

                    products_col.update_one(
                        {"_id": prod_id},
                        {
                            "$set": update_fields,
                            "$inc": {
                                "current_stock": stock_delta,
                                "total_qty_required": qty,
                                "pending_requirements": pending_delta,
                            }
                        }
                    )
                else:
                    new_products_cnt += 1
                    initial_stock = qty if order_status.lower() == "fulfilled" else max(qty, 5)
                    pending_qty = qty if order_status.lower() in ("pending", "approved") else 0
                    
                    new_product_doc = {
                        "name": clean_name,
                        "normalized_name": norm_name,
                        "details": details,
                        "aliases": aliases,
                        "keywords": keywords,
                        "category": category,
                        "sub_category": sub_category,
                        "current_stock": initial_stock,
                        "min_stock": min_stock_val,
                        "unit_price": unit_price if unit_price > 0 else (amount / qty if qty else 0),
                        "supplier": vendor,
                        "market": market,
                        "status": "low_stock" if initial_stock <= min_stock_val else "in_stock",
                        "total_expense": amount if order_status.lower() == "fulfilled" else 0,
                        "total_qty_purchased": qty if order_status.lower() == "fulfilled" else 0,
                        "total_qty_required": qty,
                        "pending_requirements": pending_qty,
                        "source_type": "excel_import",
                        "created_by_import_id": import_batch_id,
                        "import_batch_ids": [import_batch_id],
                        "source_files": [filename],
                        "created_at": datetime.utcnow(),
                        "updated_at": datetime.utcnow(),
                    }
                    prod_id = products_col.insert_one(new_product_doc).inserted_id

                # Insert Procurement Record with metadata
                proc_record = {
                    "product_id": prod_id,
                    "product_name": clean_name,
                    "normalized_name": norm_name,
                    "category": category,
                    "sub_category": sub_category,
                    "details": details,
                    "quantity": qty,
                    "unit_price": unit_price,
                    "amount": amount,
                    "market": market,
                    "order_date": dt_obj,
                    "order_date_str": date_str,
                    "order_status": order_status,
                    "vendor_name": vendor,
                    "approved_by": approved_by,
                    "remarks": remarks,
                    "requirement_issued_by": issued_by,
                    "url": url,
                    "source_type": "excel_import",
                    "source_file": filename,
                    "source_sheet": sheet_name,
                    "import_id": import_id,
                    "import_batch_id": import_batch_id,
                    "row_hash": row_hash,
                    "created_at": datetime.utcnow(),
                }
                procurement_col.insert_one(proc_record)

        # Update stock status for all products
        for prod in products_col.find({}):
            cur = prod.get("current_stock", 0)
            mins = prod.get("min_stock", 5)
            status_str = "out_of_stock" if cur == 0 else ("low_stock" if cur <= mins else "in_stock")
            products_col.update_one({"_id": prod["_id"]}, {"$set": {"status": status_str}})

        # Finalize import record
        imports_col.update_one(
            {"_id": import_id},
            {
                "$set": {
                    "total_rows": total_rows,
                    "valid_records": valid_records_cnt,
                    "new_records": new_products_cnt,
                    "updated_records": updated_products_cnt,
                    "duplicate_records": duplicate_rows_cnt,
                    "rejected_records": total_rows - (valid_records_cnt + duplicate_rows_cnt),
                    "errors": errors,
                    "status": "completed",
                }
            }
        )

        return {
            "success": True,
            "filename": filename,
            "file_type": "procurement",
            "import_batch_id": import_batch_id,
            "total_rows": total_rows,
            "valid_records": valid_records_cnt,
            "new_records": new_products_cnt,
            "updated_records": updated_products_cnt,
            "duplicate_records": duplicate_rows_cnt,
            "errors": errors,
            "import_id": str(import_id)
        }

    except Exception as e:
        logger.exception("Error in process_procurement_data: %s", e)
        imports_col.update_one(
            {"_id": import_id},
            {"$set": {"status": "failed", "errors": [str(e)]}}
        )
        raise e


def process_expenses_data(
    file_bytes: bytes,
    filename: str,
    uploaded_by: str = "admin"
) -> Dict[str, Any]:
    """
    Process Expenses Workbook.
    Assigns unique import_batch_id, retains metadata, ensures idempotency.
    """
    init_mongo_indexes()
    products_col = get_products_collection()
    expenses_col = get_expenses_collection()
    imports_col = get_imports_collection()

    sheets = read_excel_sheets(file_bytes)

    import_batch_id = f"batch_exp_{uuid.uuid4().hex[:12]}"

    import_record = {
        "import_batch_id": import_batch_id,
        "filename": filename,
        "file_type": "expenses",
        "source_type": "excel_import",
        "upload_timestamp": datetime.utcnow(),
        "uploaded_by": uploaded_by,
        "total_rows": 0,
        "valid_records": 0,
        "new_records": 0,
        "updated_records": 0,
        "duplicate_records": 0,
        "rejected_records": 0,
        "errors": [],
        "status": "in_progress",
    }
    import_id = imports_col.insert_one(import_record).inserted_id

    total_rows = 0
    new_products_cnt = 0
    updated_products_cnt = 0
    duplicate_rows_cnt = 0
    valid_records_cnt = 0
    errors = []

    try:
        for sheet_name, rows in sheets.items():
            if is_dummy_sheet(sheet_name, rows):
                logger.info("Skipping non-expense tab: %s", sheet_name)
                continue

            if not rows or len(rows) < 2:
                continue

            header_idx, col_map = find_header_row(rows)
            if not col_map:
                continue

            # Resolve all column positions flexibly
            cols = resolve_column_indices(col_map)
            comp_idx = cols["name"]
            id_idx = cols["id"]
            cat_idx = cols["category"]
            subcat_idx = cols["sub_category"]
            qty_idx = cols["qty"]
            unit_price_idx = cols["price"]
            date_idx = cols["date"]
            amount_idx = cols["amount"]
            min_stock_idx = cols["min_stock"]
            status_idx = cols["status"]
            remark_idx = cols["remarks"] if cols["remarks"] is not None else cols["details"]
            vendor_idx = cols["vendor"]

            # Process expense rows
            for r_idx in range(header_idx + 1, len(rows)):
                row = rows[r_idx]
                if not row or not any(row):
                    continue

                total_rows += 1
                raw_comp = row[comp_idx] if comp_idx is not None and comp_idx < len(row) else None
                if not raw_comp or str(raw_comp).strip() == "":
                    continue

                clean_name = clean_display_name(raw_comp)
                norm_name = normalize_text(clean_name)
                if not norm_name:
                    continue

                row_cat = str(row[cat_idx]).strip() if cat_idx is not None and cat_idx < len(row) and row[cat_idx] is not None else None
                row_subcat = str(row[subcat_idx]).strip() if subcat_idx is not None and subcat_idx < len(row) and row[subcat_idx] is not None else None
                vendor = str(row[vendor_idx]).strip() if vendor_idx is not None and vendor_idx < len(row) and row[vendor_idx] is not None else "Corporate Vendor"

                qty = parse_int_qty(row[qty_idx] if qty_idx is not None and qty_idx < len(row) else 1, default=1)
                unit_price = parse_numeric(row[unit_price_idx] if unit_price_idx is not None and unit_price_idx < len(row) else 0.0)
                amount = parse_numeric(row[amount_idx] if amount_idx is not None and amount_idx < len(row) else (qty * unit_price))
                min_stock_val = parse_int_qty(row[min_stock_idx] if min_stock_idx is not None and min_stock_idx < len(row) else max(3, int(qty * 0.5)), default=max(3, int(qty * 0.5)))
                
                raw_date = row[date_idx] if date_idx is not None and date_idx < len(row) else None
                dt_obj, date_str = parse_date_value(raw_date)

                status = str(row[status_idx]).strip() if status_idx is not None and status_idx < len(row) and row[status_idx] is not None else "Paid"
                remark = str(row[remark_idx]).strip() if remark_idx is not None and remark_idx < len(row) and row[remark_idx] is not None else ""

                category, sub_category = extract_category_and_subcategory(
                    filename=filename,
                    sheet_name=sheet_name,
                    item_name=clean_name,
                    details=remark,
                    row_category=row_cat,
                    row_subcategory=row_subcat,
                )

                # Business key deduplication hash
                row_hash = compute_row_hash(norm_name, qty, amount, unit_price, date_str, sheet_name, status, remark)

                # Check duplicate
                existing_exp = expenses_col.find_one({"row_hash": row_hash})
                if existing_exp:
                    duplicate_rows_cnt += 1
                    logger.debug("Duplicate expense row skipped: %s (%s)", clean_name, row_hash)
                    continue

                valid_records_cnt += 1
                keywords, aliases = extract_keywords_and_aliases(clean_name, remark)

                # Upsert into Master Inventory (products collection)
                existing_prod = products_col.find_one({"normalized_name": norm_name})

                if existing_prod:
                    updated_products_cnt += 1
                    prod_id = existing_prod["_id"]
                    
                    merged_keywords = sorted(list(set(existing_prod.get("keywords", []) + keywords)))
                    merged_aliases = sorted(list(set(existing_prod.get("aliases", []) + aliases)))
                    merged_sources = list(set(existing_prod.get("source_files", []) + [filename]))
                    merged_batches = list(set(existing_prod.get("import_batch_ids", []) + [import_batch_id]))

                    stock_delta = qty if status.lower() == "paid" else 0

                    update_fields = {
                        "updated_at": datetime.utcnow(),
                        "keywords": merged_keywords,
                        "aliases": merged_aliases,
                        "source_files": merged_sources,
                        "import_batch_ids": merged_batches,
                    }
                    if category and (not existing_prod.get("category") or existing_prod.get("category") in ("General", "General Supplies")):
                        update_fields["category"] = category
                    if sub_category and not existing_prod.get("sub_category"):
                        update_fields["sub_category"] = sub_category
                    if unit_price > 0 and (not existing_prod.get("unit_price") or existing_prod.get("unit_price") == 0.0):
                        update_fields["unit_price"] = unit_price
                    if remark and not existing_prod.get("details"):
                        update_fields["details"] = remark

                    products_col.update_one(
                        {"_id": prod_id},
                        {
                            "$set": update_fields,
                            "$inc": {
                                "current_stock": stock_delta,
                                "total_expense": amount,
                                "total_qty_purchased": qty,
                            }
                        }
                    )
                else:
                    new_products_cnt += 1
                    initial_stock = qty if status.lower() == "paid" else max(qty, 5)
                    
                    new_product_doc = {
                        "name": clean_name,
                        "normalized_name": norm_name,
                        "details": remark,
                        "aliases": aliases,
                        "keywords": keywords,
                        "category": category,
                        "sub_category": sub_category,
                        "current_stock": initial_stock,
                        "min_stock": min_stock_val,
                        "unit_price": unit_price if unit_price > 0 else (amount / qty if qty else 0),
                        "supplier": vendor,
                        "market": "Direct",
                        "status": "low_stock" if initial_stock <= min_stock_val else "in_stock",
                        "total_expense": amount,
                        "total_qty_purchased": qty,
                        "total_qty_required": 0,
                        "pending_requirements": 0,
                        "source_type": "excel_import",
                        "created_by_import_id": import_batch_id,
                        "import_batch_ids": [import_batch_id],
                        "source_files": [filename],
                        "created_at": datetime.utcnow(),
                        "updated_at": datetime.utcnow(),
                    }
                    prod_id = products_col.insert_one(new_product_doc).inserted_id

                # Insert Expense Record with metadata
                exp_record = {
                    "product_id": prod_id,
                    "product_name": clean_name,
                    "normalized_name": norm_name,
                    "category": category,
                    "sub_category": sub_category,
                    "quantity": qty,
                    "unit_price": unit_price,
                    "date": dt_obj,
                    "date_str": date_str,
                    "amount": amount,
                    "status": status,
                    "remark": remark,
                    "expense_month": sheet_name,
                    "source_type": "excel_import",
                    "source_file": filename,
                    "source_sheet": sheet_name,
                    "import_id": import_id,
                    "import_batch_id": import_batch_id,
                    "row_hash": row_hash,
                    "created_at": datetime.utcnow(),
                }
                expenses_col.insert_one(exp_record)

        # Update stock status
        for prod in products_col.find({}):
            cur = prod.get("current_stock", 0)
            mins = prod.get("min_stock", 5)
            status_str = "out_of_stock" if cur == 0 else ("low_stock" if cur <= mins else "in_stock")
            products_col.update_one({"_id": prod["_id"]}, {"$set": {"status": status_str}})

        # Finalize import record
        imports_col.update_one(
            {"_id": import_id},
            {
                "$set": {
                    "total_rows": total_rows,
                    "valid_records": valid_records_cnt,
                    "new_records": new_products_cnt,
                    "updated_records": updated_products_cnt,
                    "duplicate_records": duplicate_rows_cnt,
                    "rejected_records": total_rows - (valid_records_cnt + duplicate_rows_cnt),
                    "errors": errors,
                    "status": "completed",
                }
            }
        )

        return {
            "success": True,
            "filename": filename,
            "file_type": "expenses",
            "import_batch_id": import_batch_id,
            "total_rows": total_rows,
            "valid_records": valid_records_cnt,
            "new_records": new_products_cnt,
            "updated_records": updated_products_cnt,
            "duplicate_records": duplicate_rows_cnt,
            "errors": errors,
            "import_id": str(import_id)
        }

    except Exception as e:
        logger.exception("Error in process_expenses_data: %s", e)
        imports_col.update_one(
            {"_id": import_id},
            {"$set": {"status": "failed", "errors": [str(e)]}}
        )
        raise e


def auto_detect_and_import(file_bytes: bytes, filename: str, uploaded_by: str = "admin") -> Dict[str, Any]:
    """
    Intelligently detect whether an uploaded file is Procurement or Expenses and process it.
    """
    sheets = read_excel_sheets(file_bytes)
    sheet_names_lower = [s.lower() for s in sheets.keys()]

    if any("expense" in s for s in sheet_names_lower):
        return process_expenses_data(file_bytes, filename, uploaded_by)

    for name, rows in sheets.items():
        if not rows or len(rows) < 2:
            continue
        _, col_map = find_header_row(rows)
        if any("approved" in c or "vendor" in c or "market" in c for c in col_map.keys()):
            return process_procurement_data(file_bytes, filename, uploaded_by)

    first_sheet = next(iter(sheets.values()))
    _, col_map = find_header_row(first_sheet)
    if "amount" in col_map and "unit price" in col_map:
        return process_expenses_data(file_bytes, filename, uploaded_by)

    return process_procurement_data(file_bytes, filename, uploaded_by)
