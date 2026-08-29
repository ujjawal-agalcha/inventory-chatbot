import io
import logging
from datetime import datetime
from typing import Optional, List, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.mongo_inventory_service import get_all_products

logger = logging.getLogger("excel_export")


# ============================================================
# EXCEL COLOR & STYLE DEFINITIONS
# ============================================================

# Header styles
HEADER_FILL = PatternFill(start_color="FF1E293B", end_color="FF1E293B", fill_type="solid")
HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Data row styles
FONT_BASE = Font(name="Segoe UI", size=10.5)
FONT_BOLD = Font(name="Segoe UI", size=10.5, bold=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# Subtle borders
THIN_BORDER = Border(
    left=Side(style="thin", color="FFE2E8F0"),
    right=Side(style="thin", color="FFE2E8F0"),
    top=Side(style="thin", color="FFE2E8F0"),
    bottom=Side(style="thin", color="FFE2E8F0"),
)

# Alternating row background
ZEBRA_FILL = PatternFill(start_color="FFF8FAFC", end_color="FFF8FAFC", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

# ============================================================
# STOCK CONDITIONAL FORMATTING COLORS
# Thresholds:
#   Stock < 15  -> RED
#   Stock = 15  -> YELLOW
#   Stock > 15  -> GREEN
# ============================================================
STOCK_RED_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
STOCK_RED_FONT = Font(name="Segoe UI", size=10.5, bold=True, color="FF9C0006")

STOCK_YELLOW_FILL = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")
STOCK_YELLOW_FONT = Font(name="Segoe UI", size=10.5, bold=True, color="FF9C6500")

STOCK_GREEN_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
STOCK_GREEN_FONT = Font(name="Segoe UI", size=10.5, bold=True, color="FF006100")


def get_stock_style(stock_value: int):
    """
    Return appropriate (PatternFill, Font) for stock level according to exact rules:
    - Stock < 15  -> RED
    - Stock = 15  -> YELLOW
    - Stock > 15  -> GREEN
    """
    if stock_value < 15:
        return STOCK_RED_FILL, STOCK_RED_FONT
    elif stock_value == 15:
        return STOCK_YELLOW_FILL, STOCK_YELLOW_FONT
    else:
        return STOCK_GREEN_FILL, STOCK_GREEN_FONT


def format_status_label(status_code: str) -> str:
    """Format machine status string into human-readable label."""
    mapping = {
        "in_stock": "In Stock",
        "low_stock": "Low Stock",
        "out_of_stock": "Out of Stock",
    }
    return mapping.get(status_code, str(status_code).replace("_", " ").title())


# ============================================================
# MASTER SHEET EXCEL GENERATOR
# ============================================================

def generate_master_sheet_bytes() -> bytes:
    """
    Fetch the latest live inventory directly from MongoDB and generate
    a professionally styled Excel workbook (.xlsx).

    MongoDB is the SINGLE SOURCE OF TRUTH.
    """
    # 1. Fetch current inventory from MongoDB
    products: List[Dict[str, Any]] = get_all_products()
    logger.info("Generating Master Sheet for %d products from MongoDB", len(products))

    # 2. Create workbook and active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Inventory"

    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True

    # 3. Define Headers
    headers = [
        "Component / Product Name",
        "Category",
        "Price (₹)",
        "Stock Left",
        "Minimum Stock Level",
        "Supplier",
        "Status",
        "Last Updated",
    ]

    # Write Header Row
    ws.row_dimensions[1].height = 28
    for col_idx, header_title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # 4. Populate Data Rows
    for row_idx, product in enumerate(products, start=2):
        ws.row_dimensions[row_idx].height = 22
        is_even = (row_idx % 2 == 0)
        default_row_fill = WHITE_FILL if is_even else ZEBRA_FILL

        # Extract product data from MongoDB dict
        name = product.get("name", "Unknown Item")
        category = product.get("category", "General")
        unit_price = float(product.get("unit_price", 0.0))
        stock_left = int(product.get("stock", product.get("current_stock", 0)))
        min_stock = int(product.get("min_stock", 0))
        supplier = product.get("supplier", "Standard Vendor")
        status = format_status_label(product.get("status", "in_stock"))
        last_updated_raw = product.get("last_updated", "")
        
        # Format date for display
        last_updated_str = ""
        if last_updated_raw:
            try:
                if "T" in str(last_updated_raw):
                    dt = datetime.fromisoformat(str(last_updated_raw).replace("Z", "+00:00"))
                    last_updated_str = dt.strftime("%Y-%m-%d %H:%M")
                else:
                    last_updated_str = str(last_updated_raw)[:16]
            except Exception:
                last_updated_str = str(last_updated_raw)
        if not last_updated_str:
            last_updated_str = datetime.utcnow().strftime("%Y-%m-%d %H:%M")

        # Cell 1: Component Name
        c1 = ws.cell(row=row_idx, column=1, value=name)
        c1.font = FONT_BOLD
        c1.alignment = ALIGN_LEFT
        c1.fill = default_row_fill
        c1.border = THIN_BORDER

        # Cell 2: Category
        c2 = ws.cell(row=row_idx, column=2, value=category)
        c2.font = FONT_BASE
        c2.alignment = ALIGN_LEFT
        c2.fill = default_row_fill
        c2.border = THIN_BORDER

        # Cell 3: Price
        c3 = ws.cell(row=row_idx, column=3, value=unit_price)
        c3.font = FONT_BASE
        c3.alignment = ALIGN_RIGHT
        c3.number_format = "₹#,##0.00"
        c3.fill = default_row_fill
        c3.border = THIN_BORDER

        # Cell 4: Stock Left (WITH CONDITIONAL COLOR FORMATTING)
        stock_fill, stock_font = get_stock_style(stock_left)
        c4 = ws.cell(row=row_idx, column=4, value=stock_left)
        c4.font = stock_font
        c4.fill = stock_fill
        c4.alignment = ALIGN_CENTER
        c4.number_format = "#,##0"
        c4.border = THIN_BORDER

        # Cell 5: Minimum Stock Level
        c5 = ws.cell(row=row_idx, column=5, value=min_stock)
        c5.font = FONT_BASE
        c5.alignment = ALIGN_CENTER
        c5.number_format = "#,##0"
        c5.fill = default_row_fill
        c5.border = THIN_BORDER

        # Cell 6: Supplier
        c6 = ws.cell(row=row_idx, column=6, value=supplier)
        c6.font = FONT_BASE
        c6.alignment = ALIGN_LEFT
        c6.fill = default_row_fill
        c6.border = THIN_BORDER

        # Cell 7: Status
        c7 = ws.cell(row=row_idx, column=7, value=status)
        c7.font = FONT_BASE
        c7.alignment = ALIGN_CENTER
        c7.fill = default_row_fill
        c7.border = THIN_BORDER

        # Cell 8: Last Updated
        c8 = ws.cell(row=row_idx, column=8, value=last_updated_str)
        c8.font = FONT_BASE
        c8.alignment = ALIGN_CENTER
        c8.fill = default_row_fill
        c8.border = THIN_BORDER

    # 5. Freeze Header Row
    ws.freeze_panes = "A2"

    # 6. Enable Auto-filter across all columns and rows
    last_row = max(len(products) + 1, 1)
    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    # 7. Auto-fit column widths with readable padding
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val_str = str(cell.value or "")
            if cell.number_format and "₹" in cell.number_format and isinstance(cell.value, (int, float)):
                val_str = f"₹{cell.value:,.2f}"
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    # Specific tweaks for best layout
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 28)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width, 18)
    ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width, 15)
    ws.column_dimensions["D"].width = max(ws.column_dimensions["D"].width, 14)
    ws.column_dimensions["E"].width = max(ws.column_dimensions["E"].width, 20)
    ws.column_dimensions["F"].width = max(ws.column_dimensions["F"].width, 26)
    ws.column_dimensions["G"].width = max(ws.column_dimensions["G"].width, 16)
    ws.column_dimensions["H"].width = max(ws.column_dimensions["H"].width, 20)

    # 8. Save to memory buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
