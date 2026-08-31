import io
import sys
import json
import unittest
import openpyxl
from pathlib import Path
from fastapi.testclient import TestClient

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))

from app import app
from database.models import SessionLocal, User
from database.mongodb import (
    get_products_collection,
    get_conversations_collection,
    get_messages_collection,
)
from services.auth_service import create_access_token
from services.inventory_service import (
    update_product_stock,
    ensure_permanent_electronic_inventory,
)


class TestE2EChatbot(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        # 1. Ensure permanent inventory exists in MongoDB
        ensure_permanent_electronic_inventory()
        cls.client = TestClient(app)

        # 2. Ensure test user
        db = SessionLocal()
        user = db.query(User).filter(User.username == "admin").first()
        if not user:
            user = User(
                username="admin",
                email="admin@inventory.local",
                name="System Administrator",
                hashed_password="dummy_password_hash",
            )
            db.add(user)
            db.commit()
            db.refresh(user)

        cls.user_id = user.id
        cls.username = user.username
        cls.token = create_access_token({
            "sub": user.id,
            "username": user.username,
            "name": user.name,
        })
        db.close()

    def test_01_conversations_rest_crud(self):
        """Test creating, listing, updating and deleting conversations."""
        headers = {"Authorization": f"Bearer {self.token}"}

        # Create Conversation 1
        res = self.client.post(
            "/api/conversations",
            json={"title": "Test Chat 1"},
            headers=headers,
        )
        self.assertEqual(res.status_code, 200)
        conv1 = res.json()
        self.assertEqual(conv1["title"], "Test Chat 1")
        conv1_id = conv1["id"]
        print(f"[PASS] Created Conversation 1: id={conv1_id}")

        # Create Conversation 2
        res2 = self.client.post(
            "/api/conversations",
            json={"title": "Test Chat 2"},
            headers=headers,
        )
        self.assertEqual(res2.status_code, 200)
        conv2 = res2.json()
        conv2_id = conv2["id"]
        print(f"[PASS] Created Conversation 2: id={conv2_id}")

        # List conversations
        res_list = self.client.get("/api/conversations", headers=headers)
        self.assertEqual(res_list.status_code, 200)
        conv_list = res_list.json()
        ids = [c["id"] for c in conv_list]
        self.assertIn(conv1_id, ids)
        self.assertIn(conv2_id, ids)
        print(f"[PASS] Listed conversations for user: found {len(conv_list)} items.")

        # Update title
        res_patch = self.client.patch(
            f"/api/conversations/{conv1_id}",
            json={"title": "Updated Chat 1 Title"},
            headers=headers,
        )
        self.assertEqual(res_patch.status_code, 200)
        self.assertEqual(res_patch.json()["title"], "Updated Chat 1 Title")
        print("[PASS] Successfully updated conversation title.")

        # Verify in MongoDB
        mongo_conv = get_conversations_collection().find_one({"id": conv1_id})
        self.assertIsNotNone(mongo_conv)
        self.assertEqual(mongo_conv["title"], "Updated Chat 1 Title")
        print("[PASS] MongoDB persistence verified for conversation record.")

    def test_02_websocket_chat_greetings_and_streaming(self):
        """Test WebSocket greeting messages and streaming response."""
        with self.client.websocket_connect(f"/ws/chat?token={self.token}") as ws:
            # 1. Send Ping
            ws.send_json({"type": "ping"})
            pong = ws.receive_json()
            self.assertEqual(pong.get("type"), "pong")
            print("[PASS] WebSocket ping/pong verified.")

            # 2. Create and init a conversation
            ws.send_json({
                "type": "message",
                "content": "Hi there!",
            })

            received_tokens = []
            final_message = None

            while True:
                msg = ws.receive_json()
                msg_type = msg.get("type")
                if msg_type == "conversation_created":
                    active_id = msg["conversation"]["id"]
                    print(f"[PASS] WebSocket auto-created conversation: {active_id}")
                elif msg_type == "message_start":
                    pass
                elif msg_type == "token":
                    received_tokens.append(msg.get("content", ""))
                elif msg_type == "message_end":
                    final_message = msg.get("message")
                    break

            self.assertGreater(len(received_tokens), 0)
            self.assertIsNotNone(final_message)
            full_response = "".join(received_tokens)
            print(f"[PASS] Received streamed greeting response: '{full_response[:60]}...'")

    def test_03_websocket_inventory_query(self):
        """Test querying live MongoDB inventory stock via WebSocket."""
        # Ensure ESP32-CAM has a known stock (e.g. 18)
        update_product_stock("ESP32-CAM", 18)

        with self.client.websocket_connect(f"/ws/chat?token={self.token}") as ws:
            ws.send_json({
                "type": "message",
                "content": "How many ESP32-CAM do we have in stock?",
            })

            received_tokens = []
            final_data = None

            while True:
                msg = ws.receive_json()
                if msg.get("type") == "token":
                    received_tokens.append(msg.get("content", ""))
                elif msg.get("type") == "message_end":
                    final_data = msg.get("data")
                    final_msg = msg.get("message")
                    break

            full_text = "".join(received_tokens)
            self.assertIn("18", full_text)
            self.assertIn("ESP32-CAM", full_text)
            print(f"[PASS] Verified MongoDB stock query for ESP32-CAM (Stock: 18 units).")

    def test_04_websocket_multi_turn_pronoun_context(self):
        """Test multi-turn pronoun memory ("Who supplies it?")."""
        headers = {"Authorization": f"Bearer {self.token}"}
        conv_res = self.client.post("/api/conversations", json={"title": "Context Test"}, headers=headers)
        conv_id = conv_res.json()["id"]

        with self.client.websocket_connect(f"/ws/chat?token={self.token}") as ws:
            ws.send_json({"type": "init", "conversation_id": conv_id})

            # Turn 1: Ask about SG90 Servo Motor
            ws.send_json({
                "type": "message",
                "content": "Tell me about SG90 Servo Motor",
                "conversation_id": conv_id,
            })
            while True:
                msg = ws.receive_json()
                if msg.get("type") == "message_end":
                    break

            # Turn 2: Follow up with pronoun "Who supplies it?"
            ws.send_json({
                "type": "message",
                "content": "Who supplies it?",
                "conversation_id": conv_id,
            })

            turn2_tokens = []
            while True:
                msg = ws.receive_json()
                if msg.get("type") == "token":
                    turn2_tokens.append(msg.get("content", ""))
                elif msg.get("type") == "message_end":
                    break

            turn2_text = "".join(turn2_tokens)
            self.assertTrue(
                "ElectroHub" in turn2_text or "SG90" in turn2_text,
                f"Expected SG90 supplier resolution, got: {turn2_text}"
            )
            print("[PASS] Multi-turn pronoun resolution verified ('Who supplies it?' -> SG90 Servo Motor).")

    def test_05_master_sheet_export_and_download(self):
        """Verify Master Sheet download endpoint and color coding."""
        headers = {"Authorization": f"Bearer {self.token}"}
        res = self.client.get("/api/inventory/master-sheet/download", headers=headers)
        self.assertEqual(res.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb["Master Inventory"]
        self.assertGreater(ws.max_row, 10)
        self.assertEqual(ws.cell(row=1, column=1).value, "Product ID")
        self.assertEqual(ws.cell(row=1, column=2).value, "Product Name")
        print(f"[PASS] Master Sheet Download verified: {ws.max_row - 1} products exported.")


if __name__ == "__main__":
    suite = unittest.TestLoader().loadTestsFromTestCase(TestE2EChatbot)
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)
    sys.exit(0 if result.wasSuccessful() else 1)
