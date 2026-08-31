import io
import logging
from datetime import datetime
from typing import Optional, List, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database.repositories.product_repository import get_all_products_from_mongo

logger = logging.getLogger("excel.master_sheet")

# ============================================================
# EXCEL COLOR & STYLE DEFINITIONS
# ============================================================

# Header styles
HEADER_FILL = PatternFill(start_color="FF1E293B", end_color="FF1E293B", fill_type="solid")
HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Data row styles
FONT_BASE = Font(name="Segoe UI", size=10.5)
FONT_BOLD = Font(name="Segoe UI", size=10.5, bold=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# Subtle borders
THIN_BORDER = Border(
    left=Side(style="thin", color="FFE2E8F0"),
    right=Side(style="thin", color="FFE2E8F0"),
    top=Side(style="thin", color="FFE2E8F0"),
    bottom=Side(style="thin", color="FFE2E8F0"),
)

# Alternating row background
ZEBRA_FILL = PatternFill(start_color="FFF8FAFC", end_color="FFF8FAFC", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

# ============================================================
# STOCK CONDITIONAL FORMATTING COLORS
# Thresholds:
#   Stock < min_stock (or 15)  -> RED
#   Stock = min_stock (or 15)  -> YELLOW
#   Stock > min_stock (or 15)  -> GREEN
# ============================================================
STOCK_RED_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
STOCK_RED_FONT = Font(name="Segoe UI", size=10.5, bold=True, color="FF9C0006")

STOCK_YELLOW_FILL = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")
STOCK_YELLOW_FONT = Font(name="Segoe UI", size=10.5, bold=True, color="FF9C6500")

STOCK_GREEN_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
STOCK_GREEN_FONT = Font(name="Segoe UI", size=10.5, bold=True, color="FF006100")


def get_stock_style(stock_value: int, min_stock: int = 15):
    """
    Return appropriate (PatternFill, Font) for stock level.
    Uses the product's own min_stock threshold when available, falls back to 15.
    - Stock < threshold  -> RED
    - Stock = threshold  -> YELLOW
    - Stock > threshold  -> GREEN
    """
    threshold = min_stock if min_stock > 0 else 15
    if stock_value < threshold:
        return STOCK_RED_FILL, STOCK_RED_FONT
    elif stock_value == threshold:
        return STOCK_YELLOW_FILL, STOCK_YELLOW_FONT
    else:
        return STOCK_GREEN_FILL, STOCK_GREEN_FONT


def format_status_label(status_code: str) -> str:
    """Format machine status string into human-readable label."""
    mapping = {
        "in_stock": "In Stock",
        "low_stock": "Low Stock",
        "out_of_stock": "Out of Stock",
    }
    return mapping.get(status_code, str(status_code).replace("_", " ").title())


def generate_master_sheet_bytes() -> bytes:
    """
    Fetch the latest live inventory directly from MongoDB and generate
    a professionally styled Excel workbook (.xlsx).
    """
    products: List[Dict[str, Any]] = get_all_products_from_mongo()
    logger.info("Generating Master Sheet for %d products from MongoDB", len(products))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Inventory"

    ws.views.sheetView[0].showGridLines = True

    headers = [
        "Product ID",
        "Product Name",
        "Category",
        "Sub-category",
        "Supplier",
        "Price (₹)",
        "Purchased Qty",
        "Stock Left",
        "Minimum Stock",
        "Total Expense (₹)",
        "Pending Requirement",
        "Source File(s)",
        "Status",
        "Last Updated",
    ]

    ws.row_dimensions[1].height = 28
    for col_idx, header_title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for row_idx, product in enumerate(products, start=2):
        ws.row_dimensions[row_idx].height = 22
        is_even = (row_idx % 2 == 0)
        default_row_fill = WHITE_FILL if is_even else ZEBRA_FILL

        prod_id = str(product.get("id", ""))
        name = product.get("name", "Unknown Item")
        category = product.get("category", "General")
        sub_category = product.get("sub_category", "") or "General"
        supplier = product.get("supplier", "Standard Vendor")
        unit_price = float(product.get("unit_price", 0.0))
        purchased_qty = int(product.get("total_qty_purchased", 0))
        stock_left = int(product.get("stock", product.get("current_stock", 0)))
        min_stock = int(product.get("min_stock", 0))
        total_expense = float(product.get("total_expense", 0.0))
        pending_req = int(product.get("pending_requirements", 0))
        source_files_list = product.get("source_files", [])
        source_files_str = ", ".join(source_files_list) if source_files_list else product.get("source_type", "permanent_inventory")
        status = format_status_label(product.get("status", "in_stock"))
        last_updated_raw = product.get("last_updated", "")
        
        last_updated_str = ""
        if last_updated_raw:
            try:
                if "T" in str(last_updated_raw):
                    dt = datetime.fromisoformat(str(last_updated_raw).replace("Z", "+00:00"))
                    last_updated_str = dt.strftime("%Y-%m-%d %H:%M")
                else:
                    last_updated_str = str(last_updated_raw)[:16]
            except Exception:
                last_updated_str = str(last_updated_raw)
        if not last_updated_str:
            last_updated_str = datetime.utcnow().strftime("%Y-%m-%d %H:%M")

        # Cell 1: Product ID
        c1 = ws.cell(row=row_idx, column=1, value=prod_id)
        c1.font = FONT_BASE
        c1.alignment = ALIGN_CENTER
        c1.fill = default_row_fill
        c1.border = THIN_BORDER

        # Cell 2: Product Name
        c2 = ws.cell(row=row_idx, column=2, value=name)
        c2.font = FONT_BOLD
        c2.alignment = ALIGN_LEFT
        c2.fill = default_row_fill
        c2.border = THIN_BORDER

        # Cell 3: Category
        c3 = ws.cell(row=row_idx, column=3, value=category)
        c3.font = FONT_BASE
        c3.alignment = ALIGN_LEFT
        c3.fill = default_row_fill
        c3.border = THIN_BORDER

        # Cell 4: Sub-category
        c4 = ws.cell(row=row_idx, column=4, value=sub_category)
        c4.font = FONT_BASE
        c4.alignment = ALIGN_LEFT
        c4.fill = default_row_fill
        c4.border = THIN_BORDER

        # Cell 5: Supplier
        c5 = ws.cell(row=row_idx, column=5, value=supplier)
        c5.font = FONT_BASE
        c5.alignment = ALIGN_LEFT
        c5.fill = default_row_fill
        c5.border = THIN_BORDER

        # Cell 6: Price
        c6 = ws.cell(row=row_idx, column=6, value=unit_price)
        c6.font = FONT_BASE
        c6.alignment = ALIGN_RIGHT
        c6.number_format = "₹#,##0.00"
        c6.fill = default_row_fill
        c6.border = THIN_BORDER

        # Cell 7: Purchased Qty
        c7 = ws.cell(row=row_idx, column=7, value=purchased_qty)
        c7.font = FONT_BASE
        c7.alignment = ALIGN_CENTER
        c7.number_format = "#,##0"
        c7.fill = default_row_fill
        c7.border = THIN_BORDER

        # Cell 8: Stock Left
        stock_fill, stock_font = get_stock_style(stock_left, min_stock)
        c8 = ws.cell(row=row_idx, column=8, value=stock_left)
        c8.font = stock_font
        c8.fill = stock_fill
        c8.alignment = ALIGN_CENTER
        c8.number_format = "#,##0"
        c8.border = THIN_BORDER

        # Cell 9: Minimum Stock Level
        c9 = ws.cell(row=row_idx, column=9, value=min_stock)
        c9.font = FONT_BASE
        c9.alignment = ALIGN_CENTER
        c9.number_format = "#,##0"
        c9.fill = default_row_fill
        c9.border = THIN_BORDER

        # Cell 10: Total Expense
        c10 = ws.cell(row=row_idx, column=10, value=total_expense)
        c10.font = FONT_BASE
        c10.alignment = ALIGN_RIGHT
        c10.number_format = "₹#,##0.00"
        c10.fill = default_row_fill
        c10.border = THIN_BORDER

        # Cell 11: Pending Requirement
        c11 = ws.cell(row=row_idx, column=11, value=pending_req)
        c11.font = FONT_BASE
        c11.alignment = ALIGN_CENTER
        c11.number_format = "#,##0"
        c11.fill = default_row_fill
        c11.border = THIN_BORDER

        # Cell 12: Source File(s)
        c12 = ws.cell(row=row_idx, column=12, value=source_files_str)
        c12.font = FONT_BASE
        c12.alignment = ALIGN_LEFT
        c12.fill = default_row_fill
        c12.border = THIN_BORDER

        # Cell 13: Status
        c13 = ws.cell(row=row_idx, column=13, value=status)
        c13.font = FONT_BASE
        c13.alignment = ALIGN_CENTER
        c13.fill = default_row_fill
        c13.border = THIN_BORDER

        # Cell 14: Last Updated
        c14 = ws.cell(row=row_idx, column=14, value=last_updated_str)
        c14.font = FONT_BASE
        c14.alignment = ALIGN_CENTER
        c14.fill = default_row_fill
        c14.border = THIN_BORDER

    ws.freeze_panes = "A2"

    last_row = max(len(products) + 1, 1)
    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val_str = str(cell.value or "")
            if cell.number_format and "₹" in cell.number_format and isinstance(cell.value, (int, float)):
                val_str = f"₹{cell.value:,.2f}"
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 24)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width, 28)
    ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width, 18)
    ws.column_dimensions["D"].width = max(ws.column_dimensions["D"].width, 18)
    ws.column_dimensions["E"].width = max(ws.column_dimensions["E"].width, 22)
    ws.column_dimensions["F"].width = max(ws.column_dimensions["F"].width, 15)
    ws.column_dimensions["G"].width = max(ws.column_dimensions["G"].width, 15)
    ws.column_dimensions["H"].width = max(ws.column_dimensions["H"].width, 14)
    ws.column_dimensions["I"].width = max(ws.column_dimensions["I"].width, 16)
    ws.column_dimensions["J"].width = max(ws.column_dimensions["J"].width, 18)
    ws.column_dimensions["K"].width = max(ws.column_dimensions["K"].width, 20)
    ws.column_dimensions["L"].width = max(ws.column_dimensions["L"].width, 24)
    ws.column_dimensions["M"].width = max(ws.column_dimensions["M"].width, 16)
    ws.column_dimensions["N"].width = max(ws.column_dimensions["N"].width, 20)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
