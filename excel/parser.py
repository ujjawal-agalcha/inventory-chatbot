import io
import re
import logging
from datetime import datetime
from typing import Dict, Any, List, Tuple, Optional
import openpyxl

from excel.normalizer import normalize_text

logger = logging.getLogger("excel.parser")


def parse_numeric(val: Any, default: float = 0.0) -> float:
    """Safely convert strings, currency symbols, floats to float."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    s = re.sub(r"[₹$,\s]", "", s)
    try:
        return float(s)
    except (ValueError, TypeError):
        return default


def parse_int_qty(val: Any, default: int = 0) -> int:
    """Safely convert quantity to int."""
    num = parse_numeric(val, float(default))
    return int(round(num))


def parse_date_value(val: Any) -> Tuple[datetime, str]:
    """Parse date cell into datetime object and ISO string."""
    if val is None or val == "":
        now = datetime.utcnow()
        return now, now.strftime("%Y-%m-%d")
    if isinstance(val, datetime):
        return val, val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt, dt.strftime("%Y-%m-%d")
        except ValueError:
            pass
    now = datetime.utcnow()
    return now, now.strftime("%Y-%m-%d")


def is_dummy_sheet(sheet_name: str, sheet_values: List[List[Any]]) -> bool:
    """Detect if sheet is a dummy/sample/overview tab to ignore."""
    name_lower = sheet_name.lower()
    dummy_names = {"overview", "dashboard", "vendor list", "vendor contacts", "annual budget", "instructions", "sheet3"}
    if name_lower in dummy_names:
        return True
    
    for row in sheet_values[:4]:
        for cell in row:
            if cell and ("not connected to the tracker" in str(cell).lower() or "dummy tab" in str(cell).lower()):
                return True
    return False


def find_header_row(rows: List[List[Any]]) -> Tuple[int, Dict[str, int]]:
    """
    Find the header row index and map standard column names to indices.
    """
    for idx, row in enumerate(rows[:6]):
        header_map = {}
        for col_idx, cell in enumerate(row):
            if not cell:
                continue
            c_str = normalize_text(cell)
            if any(k in c_str for k in ["s no", "sno", "name", "component", "components", "product", "item", "qty", "quantity", "amount", "price", "unit price", "date", "status", "category", "supplier", "vendor"]):
                header_map[c_str] = col_idx

        if len(header_map) >= 2:
            return idx, header_map

    return 0, {}


def resolve_column_indices(col_map: Dict[str, int]) -> Dict[str, Optional[int]]:
    """Resolve column indices for any flexible variations in column headers."""
    resolved = {
        "id": None,
        "name": None,
        "category": None,
        "sub_category": None,
        "details": None,
        "qty": None,
        "amount": None,
        "price": None,
        "min_stock": None,
        "market": None,
        "date": None,
        "status": None,
        "vendor": None,
        "approved_by": None,
        "remarks": None,
        "issued_by": None,
        "url": None,
    }

    for col_name, idx in col_map.items():
        cn = col_name.lower().strip()
        
        # Product ID
        if any(k in cn for k in ["product id", "product_id", "prod id", "item id", "sku", "part number", "part no", "code"]) and not any(k in cn for k in ["name", "vendor", "supplier"]):
            if resolved["id"] is None: resolved["id"] = idx
        # Name
        elif any(k in cn for k in ["product name", "component name", "item name", "product", "component", "components", "item", "items", "part name", "description"]) and not any(k in cn for k in ["vendor", "supplier", "approv"]):
            if resolved["name"] is None: resolved["name"] = idx
        # Sub-category
        elif any(k in cn for k in ["sub category", "subcategory", "sub-category"]):
            if resolved["sub_category"] is None: resolved["sub_category"] = idx
        # Category
        elif any(k in cn for k in ["category", "division", "section", "type"]):
            if resolved["category"] is None: resolved["category"] = idx
        # Min Stock
        elif any(k in cn for k in ["minimum stock", "min stock", "min level", "threshold", "reorder level"]):
            if resolved["min_stock"] is None: resolved["min_stock"] = idx
        # Quantity
        elif any(k in cn for k in ["qty", "quantity", "units", "count", "pieces", "pcs", "stock", "current stock", "balance"]):
            if resolved["qty"] is None: resolved["qty"] = idx
        # Amount
        elif any(k in cn for k in ["amount", "total amount", "total cost", "total price", "expense", "total"]):
            if resolved["amount"] is None: resolved["amount"] = idx
        # Unit Price
        elif any(k in cn for k in ["unit price", "price", "cost", "rate", "price per unit"]):
            if resolved["price"] is None: resolved["price"] = idx
        # Market
        elif any(k in cn for k in ["market", "store"]):
            if resolved["market"] is None: resolved["market"] = idx
        # Date
        elif any(k in cn for k in ["date", "order date", "purchase date", "invoice date"]):
            if resolved["date"] is None: resolved["date"] = idx
        # Status
        elif any(k in cn for k in ["status", "order status", "state"]):
            if resolved["status"] is None: resolved["status"] = idx
        # Vendor / Supplier
        elif any(k in cn for k in ["vendor", "supplier", "dealer", "distributor"]):
            if resolved["vendor"] is None: resolved["vendor"] = idx
        # Approved by
        elif any(k in cn for k in ["approved by", "approver", "approved"]):
            if resolved["approved_by"] is None: resolved["approved_by"] = idx
        # Remarks
        elif any(k in cn for k in ["remark", "remarks", "notes", "comment"]):
            if resolved["remarks"] is None: resolved["remarks"] = idx
        # Details
        elif any(k in cn for k in ["detail", "details", "spec", "specs", "specification"]):
            if resolved["details"] is None: resolved["details"] = idx
        # Issued by
        elif any(k in cn for k in ["issued by", "requested by", "requester"]):
            if resolved["issued_by"] is None: resolved["issued_by"] = idx
        # URL
        elif "url" in cn or "link" in cn:
            if resolved["url"] is None: resolved["url"] = idx

    return resolved


def read_excel_sheets(file_bytes: bytes) -> Dict[str, List[List[Any]]]:
    """Read all sheets from Excel file bytes using openpyxl."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheets_data = {}
    for name in wb.sheetnames:
        ws = wb[name]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        sheets_data[name] = data
    return sheets_data
